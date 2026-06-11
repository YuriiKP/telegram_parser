import asyncio
import io
import json
import logging
import os
import random
from shutil import ExecError
from typing import Dict, List, Optional, Union
from venv import logger

import openpyxl
from openpyxl.styles import Font
from telethon import TelegramClient
from telethon.errors import (
    ChannelsTooMuchError,
    FloodWaitError,
    InviteHashExpiredError,
    InviteHashInvalidError,
    UserAlreadyParticipantError,
    UsernameNotOccupiedError,
)
from telethon.tl.functions.channels import GetFullChannelRequest, JoinChannelRequest
from telethon.tl.functions.messages import ImportChatInviteRequest
from telethon.tl.functions.users import GetFullUserRequest
from telethon.tl.types import (
    Channel,
    Chat,
    Message,
    PeerChannel,
    PeerUser,
    User,
    UserFull,
    UserStatusEmpty,
    UserStatusLastMonth,
    UserStatusLastWeek,
    UserStatusOffline,
    UserStatusOnline,
    UserStatusRecently,
)

from errors import JoinRequestSentError
from loader import bot
from storage import DB_M, ParsingTaskStatus, SystemAccountStatus


class TelegramParser:
    """
    Класс для парсинга участников чатов и каналов Telegram.
    Использует библиотеку opentele для работы с Telegram API.
    """

    def __init__(
        self,
        session_path: str,
        config_path: str = None,
        db_manager: DB_M = None,
        task_id: int = None,
        parse_only_active: bool = False,
        collect_bio: bool = False,
    ):
        """
        Инициализация клиента Telegram.

        Args:
            session_path (str): Путь к файлу сессии .session.
            config_path (str, optional): Путь к JSON файлу с api_id и api_hash. Defaults to None.
            db_manager: Менеджер БД для обновления статуса аккаунта при ошибках.
            task_id (int): ID задачи парсинга для проверки отмены.
            parse_only_active (bool): Флаг парсинга только активных пользователей. Defaults to False.
            collect_bio (bool): Флаг сбора биографии пользователей (поле "о себе"). Defaults to False.
        """
        self.session_path = session_path
        self.config_path = config_path
        self.db_manager = db_manager
        self.task_id = task_id
        self.parse_only_active = parse_only_active
        self.collect_bio = collect_bio

        self.client = None
        self.logger = logging.getLogger(__name__)
        self.count = 0

    async def __aenter__(self):
        await self.connect()
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        if self.client:
            await self.client.disconnect()
        if exc_type:
            print(f"Завершено с ошибкой: {exc_val}")

    async def connect(self):
        """
        Подключение к Telegram API.
        При ошибке авторизации (требуется номер телефона) помечает аккаунт как AUTH_REQUIRED.
        """
        try:
            api_id = 1
            api_hash = "пусто"

            # Если config_path предоставлен, читаем конфигурацию из JSON файла
            if self.config_path:
                with open(self.config_path, "r", encoding="utf-8") as f:
                    config = json.load(f)

                api_id = config.get("app_id") or config.get("api_id")
                api_hash = config.get("app_hash") or config.get("api_hash")
                device_model = config.get("device_model", None) or config.get(
                    "device", None
                )
                system_version = config.get("system_version", None) or config.get(
                    "sdk", None
                )
                lang_code = config.get("lang_code", None)
                system_lang_code = (
                    config.get("system_lang_code", None)
                    or config.get("system_lang_pack", None)
                    or "en-US"
                )
                app_version = config.get("app_version", None)

                if not api_id or not api_hash:
                    raise ValueError("В JSON файле отсутствуют api_id или api_hash")

            self.client = TelegramClient(
                session=self.session_path,
                api_id=int(api_id) if api_id else None,
                api_hash=api_hash,
                #
                device_model=device_model,
                system_version=system_version,
                lang_code=lang_code,
                system_lang_code=system_lang_code,
                app_version=app_version,
                #
                connection_retries=3,
                timeout=9,
                raise_last_call_error=True,
            )

            # Подключаемся без авторизации (не вызываем start())
            await self.client.connect()

            # Проверяем, авторизован ли пользователь
            if not await self.client.is_user_authorized():
                self.logger.error(f"Аккаунт не авторизован: {self.session_path}")
                if self.db_manager:
                    await self.db_manager.set_system_account_status(
                        self.session_path, SystemAccountStatus.AUTH_REQUIRED
                    )
                raise Exception(f"Аккаунт не авторизован: {self.session_path}")

            self.logger.info("Успешное подключение к Telegram API")

        except FileNotFoundError:
            self.logger.error(f"Файл конфигурации не найден: {self.config_path}")
            raise

        except json.JSONDecodeError:
            self.logger.error(f"Ошибка чтения JSON файла: {self.config_path}")
            raise

        except FloodWaitError as e:
            self.logger.error(f"FloodWait ошибка для аккаунта {self.session_path}: {e}")
            if self.db_manager:
                await self.db_manager.set_system_account_status(
                    self.session_path, SystemAccountStatus.FLOOD_WAIT
                )
            raise

        except Exception as e:
            self.logger.error(
                f"Ошибка подключения для аккаунта {self.session_path}: {e}"
            )
            # Проверяем, является ли ошибка связанной с авторизацией
            error_str = str(e).lower()

            if any(
                keyword in error_str
                for keyword in [
                    "phone",
                    "auth",
                    "sign in",
                    "login",
                    "session",
                    "not authorized",
                ]
            ):
                if self.db_manager:
                    await self.db_manager.set_system_account_status(
                        self.session_path, SystemAccountStatus.AUTH_REQUIRED
                    )
                raise
            else:
                if self.db_manager:
                    await self.db_manager.set_system_account_status(
                        self.session_path, SystemAccountStatus.BANNED
                    )
                raise

    async def disconnect(self):
        """
        Отключение от Telegram API.
        """
        if self.client:
            await self.client.disconnect()
            self.logger.info("Отключение от Telegram API")

    async def check_cancelled(self):
        """Проверить, отменена ли задача (если task_id задан)"""
        if self.task_id is None or self.db_manager is None:
            return False

        self.count += 1
        if self.count % 10 == 0:
            task = await self.db_manager.get_parsing_task(self.task_id)
            if task and task.status == ParsingTaskStatus.CANCELLED:
                raise asyncio.CancelledError(
                    f"Задача {self.task_id} отменена пользователем"
                )

    async def _handle_account_error(self, error: Exception):
        """
        Обработать ошибку аккаунта и установить соответствующий статус в БД.

        Args:
            error: Исключение, которое произошло при работе с аккаунтом.
        """
        if not self.db_manager:
            return

        error_str = str(error).lower()

        # FloodWaitError
        if isinstance(error, FloodWaitError):
            await self.db_manager.set_system_account_status(
                self.session_path, SystemAccountStatus.FLOOD_WAIT
            )
            return

        # Ошибки авторизации
        if any(
            keyword in error_str
            for keyword in [
                "phone",
                "auth",
                "sign in",
                "login",
                "session",
                "not authorized",
            ]
        ):
            await self.db_manager.set_system_account_status(
                self.session_path, SystemAccountStatus.AUTH_REQUIRED
            )
            return

        # Другие ошибки, которые могут указывать на бан
        if any(
            keyword in error_str
            for keyword in ["banned", "blocked", "suspended", "deleted"]
        ):
            await self.db_manager.set_system_account_status(
                self.session_path, SystemAccountStatus.BANNED
            )
            return

        # По умолчанию помечаем как OK
        await self.db_manager.set_system_account_status(
            self.session_path, SystemAccountStatus.OK
        )

    async def parse_users_chat(self, chat: str, limit: int = 10000) -> List[Dict]:
        """
        Парсинг участников чата (только если участники чата открыты).

        Args:
            chat (str): ссылка или id чата.
            limit (int): Максимальное количество участников для парсинга.

        Returns:
            List[Dict]: Список участников с их данными.
        """

        users_obj = []
        try:
            ################################################################################
            # c = 0
            ################################################################################
            async for user in self.client.iter_participants(chat, limit=limit):
                if user.bot:
                    continue

                # Проверяем отмену
                await self.check_cancelled()

                # Проверяем, нужно ли собирать пользователя (активность)
                if not self._should_collect_user(user):
                    continue

                users_obj.append(user)

                # Анти-флуд задержка, каждые 3000 делаем перерыв
                if len(users_obj) % 3000 == 0:
                    sleep = random.uniform(17.0, 19.0)
                    await asyncio.sleep(sleep)
                    self.logger.info(
                        f"Задача {self.task_id} | Сбор участников {len(users_obj)} | Перерыв {sleep}с"
                    )

            ################################################################################
            # c += 1
            # if c == 20:
            #     break
            ################################################################################

            await asyncio.sleep(random.uniform(17.0, 19.0))

            users_data = []
            for user_obj in users_obj:
                # Проверяем отмену
                await self.check_cancelled()

                # Сбор bio, если нужно
                if self.collect_bio:
                    await asyncio.sleep(random.uniform(0.7, 0.8))
                    # Получаем полную информацию о пользователе, чтобы достать 'о себе'
                    full_user = await self.client(GetFullUserRequest(user_obj))
                    user_data = await self._extract_user_data(full_user)

                    if len(users_data) % 100 == 0:
                        self.logger.info(
                            f"Задача {self.task_id} | Сбор данных {len(users_data)}"
                        )
                else:
                    # Используем только базовую информацию
                    user_data = await self._extract_user_data(user_obj)
                users_data.append(user_data)

        except asyncio.CancelledError:
            self.logger.info(f"Парсинг задачи {self.task_id} отменен")
            raise
        except ValueError as e:
            # Вступаем в чат, немного ждем и парсим
            self.logger.error(
                f"ValueError ошибка, пробую вступить в чат и повторить попытку: {e}"
            )
            await self._join_chat(chat)
            await asyncio.sleep(10)
            return await self.parse_users_chat(chat)
        except FloodWaitError as e:
            self.logger.error(f"FloodWait ошибка: {e}, {e.seconds}")
            await self._handle_account_error(e)
            raise
        except Exception as e:
            self.logger.error(f"Ошибка парсинга участников из истории: {e}")
            await self._handle_account_error(e)
            raise

        return users_data

    async def parse_users_from_history(self, chat: str, limit: int = 50) -> List[Dict]:
        """
        Парсинг участников из истории сообщений чата.

        Args:
            chat (str): ссылка или id чата.
            limit (int): Максимальное количество сообщений для анализа.

        Returns:
            List[Dict]: Список участников с их данными.
        """

        users_data = []
        try:
            # Проверяем, если пользователь сунул на парсинг писавших в чат канал, вместо чата
            entity = await self.client.get_entity(chat)
            # Если это канал (broadcast=True), а не супергруппа (megagroup=True)
            if getattr(entity, "broadcast", False):
                self.logger.error(
                    f"Задача {self.task_id} | Ошибка: передан канал вместо чата."
                )
                raise Exception(
                    "Нельзя собрать 'писавших' участников из канала, используйте супергруппу."
                )

            chat_info_history = await self.client.get_messages(entity=entity, limit=0)
            logger.info(
                f"Задача {self.task_id} | В чате {entity.title} - {chat_info_history.total} сообщений"
            )

            users_obj = []
            seen_users = set()
            total_messages = 0
            async for comment in self.client.iter_messages(entity=entity, limit=limit):
                # Проверяем отмену каждые 10 сообщений
                await self.check_cancelled()
                total_messages += 1

                if (
                    isinstance(comment.from_id, PeerUser)
                    and comment.from_id.user_id not in seen_users
                ):
                    seen_users.add(comment.from_id.user_id)
                    # Здесь добавляю объекты в список, чтобы можно было дальше получить информацию, просто по id это не получится
                    users_obj.append(comment.from_id)

                await asyncio.sleep(0.05)

                if total_messages % 500 == 0:
                    self.logger.info(
                        f"Задача {self.task_id} | Сбор участников {total_messages}/{chat_info_history.total}"
                    )
                # Анти-флуд задержка, каждые 3000 делаем перерыв
                if total_messages % 3000 == 0:
                    sleep = random.uniform(18.0, 21.0)
                    await asyncio.sleep(sleep)
                    self.logger.info(
                        f"Задача {self.task_id} | Сбор участников {total_messages} | Перерыв {sleep}с"
                    )

            sleep = random.uniform(12.0, 16.0)
            self.logger.info(
                f"Задача {self.task_id} | Собрано {len(users_obj)} уникальных участников | Перерыв перед сбором информации {sleep}с"
            )
            await asyncio.sleep(sleep)
            logger.info(
                f"Задача {self.task_id} | Приступаем к сбору информации о пользователях"
            )

            for peer_user in users_obj:
                # Проверяем отмену
                await self.check_cancelled()

                # Сбор bio, если нужно
                try:
                    if self.collect_bio:
                        await asyncio.sleep(random.uniform(0.7, 0.8))
                        # Получаем полную информацию о пользователе
                        full_user = await self.client(GetFullUserRequest(peer_user))
                        user = full_user.users[0]
                        full_info = full_user.full_user
                    else:
                        await asyncio.sleep(random.uniform(0.2, 0.4))
                        # Получаем только базовую информацию
                        user = await self.client.get_entity(peer_user)
                        full_info = None
                except ValueError:
                    # Пользователь недоступен (удалён / деактивирован) — пропускаем
                    continue

                if user.bot:
                    continue

                # Проверяем, нужно ли собирать пользователя (активность)
                if not self._should_collect_user(user):
                    continue

                user_data = await self._extract_user_data(user, full_info)
                users_data.append(user_data)

                if len(users_data) % 100 == 0:
                    self.logger.info(
                        f"Задача {self.task_id} | Сбор данных {len(users_data)}/{len(users_obj)}"
                    )

        except asyncio.CancelledError:
            self.logger.info(f"Парсинг задачи {self.task_id} отменен")
            raise
        except ValueError as e:
            error_str = str(e)
            if (
                "PeerUser" in error_str
                or "No user has" in error_str
                or "Could not find the input entity" in error_str
            ):
                # Это ошибка поиска пользователя — пользователь удалён или недоступен
                self.logger.warning(
                    f"Задача {self.task_id} | Некоторые пользователи недоступны, "
                    f"продолжаем с собранными: {e}"
                )
                return users_data
            else:
                # Чат не найден — пробуем вступить
                self.logger.error(
                    f"Задача {self.task_id} | Чат не найден, пробую вступить: {e}"
                )
                await self._join_chat(chat)
                await asyncio.sleep(10)
                return await self.parse_users_from_history(chat)
        except FloodWaitError as e:
            self.logger.error(f"FloodWait ошибка: {e}, {e.seconds}")
            await self._handle_account_error(e)
            raise
        except Exception as e:
            self.logger.error(f"Ошибка парсинга участников из истории чата: {e}")
            await self._handle_account_error(e)
            raise

        return users_data

    async def parse_channel_commenters(
        self, chat: str, limit: int = 20000
    ) -> List[Dict]:
        """
        Парсинг комментаторов канала (если комментарии включены).

        Args:
            chat (str): ссылка на канал.
            limit (int): Максимальное количество комментариев для анализа.

        Returns:
            List[Dict]: Список комментаторов с их данными.
        """
        # Получаем полную информацию о канале
        try:
            full_channel = await self.client(GetFullChannelRequest(chat))
        except ValueError as e:
            # Вступаем в чат, немного ждем и парсим
            self.logger.error(
                f"ValueError ошибка, пробую вступить в чат и повторить попытку: {e}"
            )
            # Если нужно ждать одобрения админом
            await self._join_chat(chat)
            await asyncio.sleep(10)

        except Exception as e:
            self.logger.error(f"Ошибка получения информации о канале: {e}")
            await self._handle_account_error(e)
            raise

        # id привязанного чата
        chat = full_channel.full_chat.linked_chat_id
        if chat is None:
            return []
        return await self.parse_users_from_history(chat, limit)

    async def parse_channel_subscribers(
        self, chat_id: str, limit: int = 10000
    ) -> List[Dict]:
        """
        Парсинг подписчиков канала.

        Args:
            chat_id (str): ID канала или username.
            limit (int): Максимальное количество подписчиков для парсинга.

        Returns:
            List[Dict]: Список подписчиков с их данными.
        """
        # 1. Создаем пригласительную ссылку без ограничения по количеству участников
        self.logger.info(
            f"Начинаем парсинг подписчиков канала, создаем пригласительную ссылку для канала {chat_id}"
        )
        invite_link = await bot.create_chat_invite_link(
            chat_id=chat_id,
            name=f"Парсинг {self.task_id}",
            member_limit=None,  # Без ограничения по количеству участников
        )

        # 2. Вступаем в канал по пригласительной ссылке
        self.logger.info("Вступаем в канал по пригласительной ссылке")
        try:
            await self._join_chat(invite_link.invite_link)
            self.logger.info("Успешно вступили в канал")
        except Exception as e:
            # Если ссылка истекла, пробуем создать новую
            if "expired" in str(e).lower() or "not valid" in str(e).lower():
                self.logger.warning("Пригласительная ссылка истекла, создаем новую")
                invite_link = await bot.create_chat_invite_link(
                    chat_id=chat_id,
                    name=f"Парсинг {self.task_id} повторно",
                    member_limit=None,
                )
                await self._join_chat(invite_link.invite_link)
                self.logger.info("Успешно вступили в канал по новой ссылке")
            else:
                self.logger.error(f"Ошибка вступления в канал: {e}")
                raise Exception(
                    f"Не удалось вступить в канал по пригласительной ссылке: {e}"
                )

        # 3. Добавляем аккаунт в админы канала
        self.logger.info("Добавляем аккаунт в администраторы канала")
        try:
            me = await self.client.get_me()
            self.logger.info(f"ID аккаунта для назначения админом: {me.id}")

            # Назначаем аккаунт администратором с минимальными правами
            # Для парсинга достаточно прав на просмотр участников
            success = await bot.promote_chat_member(
                chat_id=chat_id,
                user_id=int(me.id),
                can_change_info=False,
                can_invite_users=True,
                can_promote_members=False,
                can_manage_chat=False,
                can_delete_messages=False,
                can_manage_video_chats=False,
                can_restrict_members=False,
                can_pin_messages=False,
                can_manage_topics=False,
                can_post_messages=False,
                can_edit_messages=False,
                can_manage_direct_messages=False,
                is_anonymous=False,
            )

            if success:
                self.logger.info(
                    f"Аккаунт {me.id} успешно назначен администратором канала"
                )
            else:
                self.logger.warning(
                    "Не удалось назначить аккаунт администратором, но продолжаем парсинг"
                )

        except Exception as e:
            self.logger.warning(
                f"Ошибка при назначении админа: {e}. Продолжаем парсинг без прав администратора."
            )
            # Не прерываем выполнение, возможно парсинг все равно будет работать

        # 4. Ждем немного для синхронизации
        await asyncio.sleep(3)

        # 5. Получаем entity канала для парсинга
        try:
            # Пробуем получить entity канала
            if chat_id.startswith("-100") and chat_id[1:].isdigit():
                # Это числовой ID канала в строковом формате
                channel_entity = await self.client.get_entity(int(chat_id))
            else:
                # Это username или другой формат
                channel_entity = await self.client.get_entity(chat_id)
        except Exception as e:
            self.logger.error(f"Ошибка получения entity канала {chat_id}: {e}")
            raise Exception(f"Не удалось получить информацию о канале: {e}")

        # 6. Выполняем парсинг участников канала
        self.logger.info(f"Начинаем парсинг участников канала {chat_id}")

        return await self.parse_users_chat(channel_entity, limit)

    def _get_last_seen(self, user_obj) -> str:
        """
        Определение времени последней активности пользователя на основе статуса.

        Args:
            user_obj: Объект пользователя из Telethon.

        Returns:
            Строка с описанием активности.
        """
        status = user_obj.status

        if isinstance(status, UserStatusOnline):
            return "В сети"
        elif isinstance(status, UserStatusOffline):
            # Здесь возвращается конкретный объект datetime
            return status.was_online.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(status, UserStatusRecently):
            return "Недавно"
        elif isinstance(status, UserStatusLastWeek):
            return "На этой неделе"
        elif isinstance(status, UserStatusLastMonth):
            return "В этом месяце"
        else:
            return "Давно/Скрыто"

    def _should_collect_user(self, user_obj) -> bool:
        """
        Определить, нужно ли собирать пользователя на основе его активности и настроек.

        Args:
            user_obj: Объект пользователя из Telethon.

        Returns:
            True, если пользователь активный (или настройка parse_only_active выключена).
        """
        # Если настройка парсинга только активных пользователей выключена, собираем всех
        if not self.parse_only_active:
            return True

        status = user_obj.status

        # Всегда собираем пользователей со статусами Online и Recently
        if isinstance(status, (UserStatusOnline, UserStatusRecently)):
            return True

        # Для статуса Offline проверяем, был ли пользователь в сети в последние 48 часов
        if isinstance(status, UserStatusOffline):
            from datetime import datetime, timezone

            if status.was_online:
                time_diff = datetime.now(timezone.utc) - status.was_online
                return time_diff.total_seconds() <= 48 * 3600  # 48 часов в секундах
            else:
                # Если время неизвестно, считаем неактивным
                return False

        # Все остальные статусы (LastWeek, LastMonth, Empty и т.д.) считаем неактивными
        return False

    async def _extract_user_data(self, user_obj, full_info=None) -> dict:
        """
        Извлечение данных пользователя.

        Args:
            user_obj: Объект пользователя из Telethon (тип User) или результат GetFullUserRequest.
            full_info: Объект расширенной информации UserFull (optional).

        Returns:
            dict: Словарь с данными пользователя.
        """
        # Обработка обратной совместимости: если передан full_user_result (имеет атрибут users)
        if hasattr(user_obj, "users") and hasattr(user_obj, "full_user"):
            # Это результат GetFullUserRequest
            full_info = user_obj.full_user
            user_obj = user_obj.users[0]

        bio = ""
        if full_info and hasattr(full_info, "about"):
            bio = full_info.about or ""

        user_data = {
            "id": user_obj.id,
            "username": user_obj.username or "",
            "first_name": user_obj.first_name or "",
            "last_name": user_obj.last_name or "",
            "phone": user_obj.phone or "",
            "premium": str(user_obj.premium),
            # "lang_code": user_obj.lang_code or "",
            "bio": bio,
            "last_seen": self._get_last_seen(user_obj),
        }
        return user_data

    async def _join_chat(self, link: str) -> None:
        """
        Вступление в приватный чат по пригласительной ссылке

        Args:
            link str: Пригласительная ссылка

        """

        try:
            # Пробуем как invite-ссылку
            if "+" in link or "joinchat" in link:
                # Извлекаем хеш
                if "/+" in link:
                    invite_hash = link.split("/+")[1]
                elif "/joinchat/" in link:
                    invite_hash = link.split("/joinchat/")[1]
                else:
                    invite_hash = link

                await self.client(ImportChatInviteRequest(invite_hash))
                self.logger.info(f"Вступили в чат @{link}!")
            else:
                raise Exception(f"Не удалось распознать ссылку: {link}")
            #     # Пробуем как username
            #     if link.startswith("@"):
            #         link = link[1:]
            #     elif "t.me/" in link:
            #         link = link.split("t.me/")[1]

            #     entity = await self.client.get_entity(link)
            #     await self.client(JoinChannelRequest(entity))
            #     self.logger.info(f"Вступили в чат @{link}!")

        except UserAlreadyParticipantError:
            self.logger.info("Вы уже состоите в этом чате.")
        except InviteHashExpiredError:
            self.logger.error("Ссылка-приглашение истекла")
            raise
        except InviteHashInvalidError:
            self.logger.error("Неверная ссылка-приглашение")
            raise
        except UsernameNotOccupiedError:
            self.logger.error("Пользователь/чат с таким username не найден")
            raise
        except ChannelsTooMuchError:
            self.logger.error("Вы вступили в слишком много каналов/групп")
            raise
        except Exception as e:
            # Если нужно ждать одобрения
            if "successfully requested to join" in str(e):
                self.logger.warning("Отправлена заявка на вступление.")
                raise JoinRequestSentError("Нужно ждать одобрения админом")

            self.logger.error(f"Критическая ошибка вступления: {e}")
            await self._handle_account_error(e)
            raise

    def get_excel_bytes(self, users_data: List[Dict]) -> io.BytesIO:
        """
        Получение данных участников в виде Excel файла в памяти.

        Args:
            users_data (List[Dict]): Список данных участников.

        Returns:
            io.BytesIO: Excel файл в памяти.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"

        # Заголовки
        headers = [
            "ID",
            "Username",
            "First Name",
            "Last Name",
            "Phone",
            "Premium",
            "Bio",
            "Last Seen",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Данные
        for row_num, user in enumerate(users_data, 2):
            sheet.cell(row=row_num, column=1, value=user["id"])
            sheet.cell(
                row=row_num,
                column=2,
                value=f"@{user['username']}" if user["username"] else "",
            )
            sheet.cell(row=row_num, column=3, value=user["first_name"])
            sheet.cell(row=row_num, column=4, value=user["last_name"])
            sheet.cell(row=row_num, column=5, value=user["phone"])
            sheet.cell(
                row=row_num, column=6, value="Да" if user["premium"] == "True" else ""
            )
            sheet.cell(row=row_num, column=7, value=user["bio"])
            sheet.cell(row=row_num, column=8, value=user["last_seen"])

        # Сохраняем в BytesIO вместо файла на диске
        excel_bytes = io.BytesIO()
        workbook.save(excel_bytes)
        excel_bytes.seek(0)

        self.logger.info("Excel данные подготовлены в памяти")
        return excel_bytes

    def get_txt_bytes(self, users_data: List[Dict]) -> io.BytesIO:
        """
        Получение данных участников в виде .txt файла в памяти.

        Args:
            users_data (List[Dict]): Список данных участников.

        Returns:
            io.BytesIO: TXT файл в памяти.
        """
        txt_content = io.StringIO()
        for user in users_data:
            if user["username"]:
                txt_content.write(f"@{user['username']}\n")

        # Конвертируем StringIO в BytesIO для отправки через Telegram
        txt_bytes = io.BytesIO(txt_content.getvalue().encode("utf-8"))
        txt_bytes.seek(0)

        self.logger.info("TXT данные подготовлены в памяти")
        return txt_bytes
